import openpyxl as xl

# opening 'student' for Read & Write operations
wb = xl.load_workbook(r'student.xlsx')  

# creating object for worksheet 'Student'
ws = wb['Student']

print("\nAccessing Entire sheet values:")
# returns cell values instead of cell objects
for row in ws.values:
    for value in row:
        print(value, end = ' ')
    print()

# Removing Columns
print("\Deleting Columns:")
columns_to_remove = ['Mark1','Mark2']
for col in range(ws.max_column, 1, -1):
    if ws.cell(1,col).value in columns_to_remove :
        ws.delete_cols(col)

print("\nAccessing Entire sheet values:")
# returns cell values instead of cell objects
for row in ws.values:
    for value in row:
        print(value, end = ' ')
    print()

# saving the changes into a new excel:
wb.save(r'student_new.xlsx')
