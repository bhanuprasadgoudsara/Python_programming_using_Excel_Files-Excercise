import openpyxl as xl

# opening 'student' for Read & Write operations
wb = xl.load_workbook(r'student.xlsx')  

# creating object for worksheet 'Student'
ws = wb['Student']

print(ws.max_row, ws.max_column)

# Fetching data from Excel:
print("\nAccessing Values row-wise:")
for row in range (1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        print(ws.cell(row, col).value, end = '|')
    print()

# Fetching data from Excel:
print("\nAccessing Values column-wise:")
for col in range (1, ws.max_column + 1):
    for row in range(1, ws.max_row + 1):
        print(ws.cell(row, col).value, end = ' ')
    print()

print("\nAccessing Values row-wise:")
all_rows = tuple(ws.rows)
for rows in all_rows:
    for col in rows:
        print(col.value, end = ' ')
    print()

print("\nAccessing Values column-wise:")
all_columns = tuple(ws.columns)
for cols in all_columns:
    for row in cols:
        print(row.value, end = ' ')
    print()

print("\nAccessing Entire sheet values:")
# returns cell values instead of cell objects
for row in ws.values:
    for value in row:
        print(value, end = ' ')
    print()

# Inserting a new column:
for row in range(2, 6):
    # ws.cell(row, 6).value = ws.cell(row,3).value + ws.cell(row,4).value + ws.cell(row,5).value
    ws.cell(row, 6).value = 0
    for col in range(3, 6):
        ws.cell(row, 6).value += ws.cell(row,col).value
    # ws.cell(row, 6).value *= 0.6
    
# Creating a New column 'Total_Mark'
ws['F1'] = 'Total_Mark'

# saving the changes
wb.save(r'student.xlsx')