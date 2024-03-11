import openpyxl as xl

# opening 'student' for Read & Write operations
wb = xl.load_workbook(r'student.xlsx')  

# creating object for worksheet 'Student'
ws = wb['Student']

def update_details(sid, m1):
    mark_col = 0
    for col in range ( 1, ws.max_column + 1):
        if ws.cell(1,col).value == 'Mark1':
            mark_col = col
    for row in range( 2, ws.max_row + 1 ):
        if ws.cell(row,1).value == sid:
            ws.cell(row,mark_col).value = m1

# Updating the Value:
update_details(101,92)

print("\nAccessing Entire sheet values:")
# returns cell values instead of cell objects
for row in ws.values:
    for value in row:
        print(value, end = ' ')
    print()
            
# saving the changes into a new excel:
wb.save(r'student.xlsx')