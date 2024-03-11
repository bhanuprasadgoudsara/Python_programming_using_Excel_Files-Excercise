import openpyxl as xl
from openpyxl import Workbook

def create_workbook(path):
    print('Creating & Saving a New Excel File')
    workbook = Workbook()                           # creates a new excel file
    workbook.save(path)                             # saves the excel file

def get_xl_details(obj):
    print('Details of the Excel: ', str(obj))
    print('Type of Workbook: ', type(obj))          
    print('Worksheet Names: ', obj.sheetnames)      # returns the list of worksheet names
    print('Worksheet Objects: ', obj.worksheets)    # returns the list of worksheet objects
    print('Active Worksheet Object: ',obj.active)   # returns the active worksheet name

def get_sheet_details(sheet1):
    print("Details of the Sheet: ", str(sheet1))
    print('Type of Worksheet: ',type(sheet1))

# creating a new excelsheet named 'student'
create_workbook("student.xlsx")  

# opening 'student' for Read & Write operations
wb = xl.load_workbook(r'student.xlsx')    

print('After Creation & Before Update')
get_xl_details(wb)
get_sheet_details(wb.active)

# creating a new sheet
wb.create_sheet('Student') 

# removing a existing sheet
wb.remove(wb['Sheet']) 

 # creating an object for the 'Student' worksheet
ws = wb['Student']   

# adding a new row as a list
ws.append(['Student_Id', 'Student_Name', 'Mark1', 'Mark2', 'Mark3'])

# adding new rows as a tuples
ws.append((101, 'David', 45, 56, 67))                                 
ws.append((102, 'John', 67, 76, 63))
ws.append((103, 'Mark', 84, 82, 93))
ws.append((104, 'Andrew', 94, 59, 69))

print('After Update')
get_xl_details(wb)
get_sheet_details(wb.active)

# saving the changes
wb.save(r'student.xlsx')